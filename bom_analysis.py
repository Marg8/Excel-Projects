"""
bom_analysis.py
Build Analysis engine: BOM explosion, build capability, and cost with
progressive date fallback.
"""
from __future__ import annotations

import math
import pandas as pd


# ── helpers ──────────────────────────────────────────────────────────────────

def _col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Return the first column name that matches any candidate (case-insensitive)."""
    lower_map = {str(c).strip().lower(): c for c in df.columns}
    for cand in candidates:
        found = lower_map.get(cand.strip().lower())
        if found is not None:
            return found
    return None


# ── 1. Std Cost with progressive date fallback ────────────────────────────────

def build_cost_lookup_dated(
    cost_df: pd.DataFrame | None,
) -> dict[str, dict]:
    """
    Returns {PN: {"cost": float, "date": str}} using the most recent
    non-zero cost per part number.

    Progressive fallback logic:
      Sort all records by update date DESC.
      For each PN, take the first row whose Std Cost > 0.
      This naturally falls back to earlier dates when the latest record
      has a zero or missing cost.

    Fields returned per PN:
      "cost"  – float, the selected Std Cost
      "date"  – str ISO-8601 date of the record used (traceability)
    """
    if cost_df is None or cost_df.empty:
        return {}

    df = cost_df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    pn_col   = _col(df, ["PN", "Material", "Part Number", "Part"])
    cost_col = _col(df, ["Std Cost", "StdCost", "Cost", "Price"])
    date_col = _col(df, ["Last update", "LastUpdate", "Update Date", "Date"])

    if pn_col is None or cost_col is None:
        return {}

    df["_pn"]   = df[pn_col].astype(str).str.strip()
    df["_cost"] = pd.to_numeric(df[cost_col], errors="coerce")
    df["_date"] = (
        pd.to_datetime(df[date_col], errors="coerce")
        if date_col
        else pd.Timestamp("1900-01-01")
    )

    # Keep only rows with a valid positive cost
    valid = df[df["_cost"].notna() & (df["_cost"] > 0)].copy()
    # Sort descending so groupby.first() picks the most recent
    valid = valid.sort_values("_date", ascending=False)

    best = valid.groupby("_pn", sort=False).first().reset_index()

    result: dict[str, dict] = {}
    for _, row in best.iterrows():
        d = row["_date"]
        result[row["_pn"]] = {
            "cost": float(row["_cost"]),
            "date": d.strftime("%Y-%m-%d") if pd.notna(d) else "",
        }
    return result


def cost_date_summary(
    cost_df: pd.DataFrame | None,
    cost_dated_lookup: dict,
) -> pd.DataFrame:
    """
    Returns a traceability table: PN, Std Cost Used, Cost Effective Date.
    Includes only the parts that appear in the dated lookup (i.e., have a cost).
    """
    if not cost_dated_lookup:
        return pd.DataFrame(columns=["Part Number", "Std Cost Used", "Cost Effective Date"])

    rows = [
        {"Part Number": pn, "Std Cost Used": info["cost"], "Cost Effective Date": info["date"]}
        for pn, info in cost_dated_lookup.items()
    ]
    return pd.DataFrame(rows).sort_values("Part Number").reset_index(drop=True)


# ── 2. BOM parsing ────────────────────────────────────────────────────────────

def parse_bom(
    bom_df: pd.DataFrame | None,
    alt_bom: int = 1,
    bom_level: int = 1,
) -> pd.DataFrame:
    """
    Returns a clean BOM DataFrame with columns:
      Material, Component, Component_Desc, BOM_Usage, BOM_Level, Alt_BOM, MRP

    BOM_Usage = qty of Component per 1 unit of Material.
        Selects one Alt BOM version per Material:
            - use `alt_bom` when that version exists for the Material,
            - otherwise fallback to the next available version for that Material.
        Then filters to the requested BOM level.
    """
    if bom_df is None or bom_df.empty:
        return pd.DataFrame()

    df = bom_df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    mat_col    = _col(df, ["Material"])
    comp_col   = _col(df, ["Component"])
    desc_col   = _col(df, ["Component Description", "Component Desc"])
    qty_col    = _col(df, ["Quantity", "Qty"])
    base_col   = _col(df, ["Base quantity", "Base Quantity", "BaseQty"])
    usage_col  = _col(df, ["Usage"])
    level_col  = _col(df, ["BOM Level", "Level"])
    altbom_col = _col(df, ["Alt BOM", "AltBOM", "Alt_BOM"])
    mrp_col    = _col(df, ["MRP"])
    mat_desc_col = _col(df, ["Material Description"])

    if mat_col is None or comp_col is None:
        return pd.DataFrame()

    out = pd.DataFrame()
    out["Material"]       = df[mat_col].astype(str).str.strip()
    out["Component"]      = df[comp_col].astype(str).str.strip()
    out["Component_Desc"] = (
        df[desc_col].astype(str).str.strip() if desc_col else out["Component"]
    )
    out["Material_Desc"]  = (
        df[mat_desc_col].astype(str).str.strip() if mat_desc_col else out["Material"]
    )
    out["BOM_Level"] = (
        pd.to_numeric(df[level_col], errors="coerce").fillna(1).astype(int)
        if level_col else 1
    )
    out["Alt_BOM"] = (
        pd.to_numeric(df[altbom_col], errors="coerce").fillna(1).astype(int)
        if altbom_col else 1
    )
    out["MRP"] = df[mrp_col].astype(str).str.strip() if mrp_col else ""

    # Compute usage per unit of parent
    if usage_col:
        out["BOM_Usage"] = pd.to_numeric(df[usage_col], errors="coerce").fillna(0.0)
    elif qty_col and base_col:
        qty  = pd.to_numeric(df[qty_col],  errors="coerce").fillna(0.0)
        base = pd.to_numeric(df[base_col], errors="coerce").replace(0, 1000.0).fillna(1000.0)
        out["BOM_Usage"] = qty / base
    elif qty_col:
        out["BOM_Usage"] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0.0)
    else:
        out["BOM_Usage"] = 1.0

    # Filters
    out = out[
        out["Material"].notna() & (out["Material"] != "nan") & (out["Material"] != "")
        & out["Component"].notna() & (out["Component"] != "nan") & (out["Component"] != "")
        & (out["BOM_Usage"] > 0)
    ]
    out = out[out["BOM_Level"] == bom_level]

    # Keep exactly one BOM version per Material.
    # Priority: requested alt_bom (default 1). If missing, use the smallest available Alt_BOM.
    chosen_alt = out.groupby("Material")["Alt_BOM"].transform(
        lambda s: alt_bom if (s == alt_bom).any() else s.min()
    )
    out = out[out["Alt_BOM"] == chosen_alt]

    # Collapse duplicate Material×Component entries by summing BOM_Usage.
    # In SAP BOM, the same component can appear in multiple operations at the
    # same level and Alt BOM — summing gives correct total usage per FG.
    out = (
        out.groupby(
            ["Material", "Component"],
            as_index=False, sort=False,
        ).agg(
            Component_Desc=("Component_Desc", "first"),
            Material_Desc=("Material_Desc",  "first"),
            BOM_Usage=("BOM_Usage",  "sum"),
            Alt_BOM=("Alt_BOM",   "first"),
            BOM_Level=("BOM_Level", "first"),
            MRP=("MRP",       "first"),
        )
    )

    return out.reset_index(drop=True)


# ── 3. BOM explosion ──────────────────────────────────────────────────────────

def explode_bom_demand(
    scheduled_df:  pd.DataFrame,
    bom_clean:     pd.DataFrame,
    fg_col:        str = "Product code",
    qty_col:       str = "ScheduledQty",
    week_col:      str = "ProductionWeek",
    week_date_col: str = "ProductionWeekDate",
) -> pd.DataFrame:
    """
    Explodes the MPS schedule through the BOM to produce component demand
    per FG per production week.

    Returns columns:
      FG, FG_Desc, Component, Component_Desc, BOM_Usage,
      ProductionWeek, ProductionWeekDate, [Line, Line Name, Quarter, Month Name],
      FG_Qty, Comp_Demand
    """
    if scheduled_df is None or scheduled_df.empty:
        return pd.DataFrame()
    if bom_clean is None or bom_clean.empty:
        return pd.DataFrame()
    if fg_col not in scheduled_df.columns:
        return pd.DataFrame()

    sch = scheduled_df.copy()
    bom = bom_clean.copy()

    sch["_fg"]  = sch[fg_col].astype(str).str.strip()
    bom["_mat"] = bom["Material"].astype(str).str.strip()

    bom_cols = ["_mat", "Component", "Component_Desc", "Material_Desc", "BOM_Usage"]
    bom_cols = [c for c in bom_cols if c in bom.columns]

    merged = sch.merge(bom[bom_cols], left_on="_fg", right_on="_mat", how="inner")
    if merged.empty:
        return pd.DataFrame()

    fg_qty_num = pd.to_numeric(merged[qty_col], errors="coerce").fillna(0)
    merged["Comp_Demand"] = fg_qty_num * merged["BOM_Usage"]

    keep: dict[str, pd.Series] = {
        "FG":             merged["_fg"],
        "FG_Desc":        merged.get("Material_Desc", merged["_fg"]),
        "Component":      merged["Component"],
        "Component_Desc": merged["Component_Desc"],
        "BOM_Usage":      merged["BOM_Usage"],
        "FG_Qty":         fg_qty_num,
        "Comp_Demand":    merged["Comp_Demand"],
    }

    # Pass through dimension columns if present
    for c in [week_col, week_date_col, "Line", "Line Name", "Quarter", "Month Name"]:
        if c in merged.columns:
            keep[c] = merged[c]

    return pd.DataFrame(keep).reset_index(drop=True)


# ── 4. Component demand pivot (raw material coverage view) ───────────────────

def component_demand_pivot(
    exploded_df: pd.DataFrame,
    week_col: str = "ProductionWeek",
) -> pd.DataFrame:
    """
    Returns a pivot: Component × Week → total demand quantity.
    Includes Component_Desc and an overall Total column.
    """
    if exploded_df is None or exploded_df.empty:
        return pd.DataFrame()
    if week_col not in exploded_df.columns:
        return pd.DataFrame()

    pvt = exploded_df.pivot_table(
        index=["Component", "Component_Desc"],
        columns=week_col,
        values="Comp_Demand",
        aggfunc="sum",
        fill_value=0,
    )
    pvt["Total"] = pvt.sum(axis=1)
    pvt = pvt.sort_values("Total", ascending=False)
    return pvt.reset_index()


# ── 5. Build capability computation ──────────────────────────────────────────

def _build_stock_lookup(stock_df: pd.DataFrame | None) -> dict[str, float]:
    if stock_df is None or stock_df.empty:
        return {}

    df = stock_df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    comp_col = _col(df, ["Component", "Material", "PN", "Part Number", "Part"])
    qty_col = _col(df, ["Stock", "Stock Qty", "Stock_Qty", "Qty", "Quantity", "On Hand"])
    if comp_col is None or qty_col is None:
        return {}

    df["_comp"] = df[comp_col].astype(str).str.strip()
    df["_qty"] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0.0)
    df = df[df["_comp"] != ""]
    return df.groupby("_comp")["_qty"].sum().to_dict()


def _build_weekly_po_table(
    rm_po_df: pd.DataFrame | None,
) -> pd.DataFrame:
    if rm_po_df is None or rm_po_df.empty:
        return pd.DataFrame(columns=["Component", "PO_WeekDate", "PO_Qty"])

    df = rm_po_df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    comp_col = _col(df, ["Component", "Material", "PN", "Part Number", "Part"])
    qty_col  = _col(df, [
        "Rec./reqd.qty", "Rec./reqd.qty2",
        "PO Qty", "Qty", "Quantity", "Open Qty", "Order Qty",
    ])
    # Del/finish = confirmed delivery date from SAP MRP element.
    # Fallback chain: Del/finish → Delivery Date → Due Date → PO Date → Date.
    date_col = _col(df, [
        "Del/finish", "Del/Finish", "Delivery Date", "Due Date",
        "PO Date", "Week Date", "Date", "Requested date",
    ])

    if comp_col is None or qty_col is None or date_col is None:
        return pd.DataFrame(columns=["Component", "PO_WeekDate", "PO_Qty"])

    out = pd.DataFrame()
    out["Component"]   = df[comp_col].astype(str).str.strip()
    out["PO_Qty"]      = pd.to_numeric(df[qty_col], errors="coerce").fillna(0.0)
    out["PO_WeekDate"] = pd.to_datetime(df[date_col], errors="coerce").dt.normalize()

    out = out[
        out["Component"].notna()
        & (out["Component"] != "")
        & out["PO_WeekDate"].notna()
        & (out["PO_Qty"] > 0)
    ]
    if out.empty:
        return pd.DataFrame(columns=["Component", "PO_WeekDate", "PO_Qty"])

    return out.groupby(["Component", "PO_WeekDate"], as_index=False)["PO_Qty"].sum()


def _cum_po_by_component_week(
    po_weekly: pd.DataFrame,
    components: pd.Series,
    week_dates: pd.Series,
) -> pd.Series:
    if po_weekly.empty:
        return pd.Series([0.0] * len(components), index=components.index, dtype="float64")

    po_map: dict[str, list[tuple[pd.Timestamp, float]]] = {}
    for _, r in po_weekly.iterrows():
        comp = str(r["Component"]).strip()
        dt = pd.to_datetime(r["PO_WeekDate"], errors="coerce")
        qty = float(pd.to_numeric(r["PO_Qty"], errors="coerce") or 0.0)
        if not comp or pd.isna(dt) or qty <= 0:
            continue
        po_map.setdefault(comp, []).append((dt.normalize(), qty))

    for comp in po_map:
        po_map[comp].sort(key=lambda t: t[0])

    values: list[float] = []
    for comp, wk in zip(components.astype(str), week_dates):
        wk_dt = pd.to_datetime(wk, errors="coerce")
        if pd.isna(wk_dt):
            values.append(0.0)
            continue
        wk_dt = wk_dt.normalize()
        cum = 0.0
        for po_dt, po_qty in po_map.get(comp.strip(), []):
            if po_dt <= wk_dt:
                cum += po_qty
            else:
                break
        values.append(cum)
    return pd.Series(values, index=components.index, dtype="float64")

def compute_build_capability(
    exploded_df: pd.DataFrame,
    stock_lookup: dict[str, float] | None = None,
    stock_df: pd.DataFrame | None = None,
    rm_po_df: pd.DataFrame | None = None,
    week_col: str = "ProductionWeek",
    week_date_col: str = "ProductionWeekDate",
) -> pd.DataFrame:
    """
    Enrich exploded demand with CTB/supply metrics.

    Supports two modes:
      1) Legacy: pass stock_lookup dict.
      2) Weekly CTB: pass stock_df and/or rm_po_df (used by app.py).
    """
    df = exploded_df.copy()

    if "Component" not in df.columns:
        return df

    # Build stock lookup from df input when provided; merge with explicit lookup.
    merged_lookup: dict[str, float] = {}
    if stock_lookup:
        merged_lookup.update({str(k).strip(): float(v) for k, v in stock_lookup.items()})
    if stock_df is not None and not stock_df.empty:
        for comp, qty in _build_stock_lookup(stock_df).items():
            merged_lookup[comp] = merged_lookup.get(comp, 0.0) + float(qty)

    df["Stock_Qty"] = df["Component"].map(merged_lookup).fillna(0.0)

    if rm_po_df is not None and not rm_po_df.empty and week_date_col in df.columns:
        po_weekly = _build_weekly_po_table(rm_po_df)
        df["Cum_PO_Qty"] = _cum_po_by_component_week(
            po_weekly,
            components=df["Component"],
            week_dates=df[week_date_col],
        )
    else:
        df["Cum_PO_Qty"] = 0.0

    df["Available_Supply"] = df["Stock_Qty"] + df["Cum_PO_Qty"]

    # Keep legacy column name used by older pipeline.
    df["Available_Qty"] = df["Available_Supply"]

    def _coverage(row):
        if row["Comp_Demand"] <= 0:
            return 100.0
        return min(100.0, row["Available_Supply"] / row["Comp_Demand"] * 100.0)

    def _buildable(row):
        if row["BOM_Usage"] <= 0:
            return row["FG_Qty"]
        return row["Available_Supply"] / row["BOM_Usage"]

    df["Coverage_Pct"]        = df.apply(_coverage, axis=1)
    df["Buildable_From_Comp"] = df.apply(_buildable, axis=1)
    df["Shortage"]            = (df["Comp_Demand"] - df["Available_Supply"]).clip(lower=0)
    # Hard CTB rule: if no stock and no PO by that week, CTB is false.
    df["CTB_Flag"] = df["Available_Supply"] > 0

    # Keep week_col referenced by callers for consistency.
    if week_col not in df.columns and week_date_col in df.columns:
        df[week_col] = pd.to_datetime(df[week_date_col], errors="coerce").dt.strftime("%Y-%m-%d")

    return df


def build_capability_summary(
    enriched_df:       pd.DataFrame,
    cost_dated_lookup: dict,
    has_stock:         bool = False,
    week_col:          str = "ProductionWeek",
    week_date_col:     str = "ProductionWeekDate",
) -> pd.DataFrame:
    """
    Summarises build capability per FG × Week.

    Returns the dashboard-ready dataset:
      Part Number, FG_Desc, Line, Line Name, Week, WeekDate, Quarter, Month Name,
      FG_Scheduled_Qty, Buildable_Qty, Material_Coverage_Pct,
      Constraint_Component, Constraint_Desc,
      Num_BOM_Components, Total_Shortage,
      Std_Cost, Cost_Effective_Date, Total_Cost
    """
    if enriched_df is None or enriched_df.empty:
        return pd.DataFrame()

    grp_keys = [c for c in
                ["FG", "FG_Desc", week_col, week_date_col,
                 "Line", "Line Name", "Quarter", "Month Name"]
                if c in enriched_df.columns]

    def _agg(g: pd.DataFrame) -> pd.Series:
        fg_qty = float(g["FG_Qty"].iloc[0])

        if has_stock and "Coverage_Pct" in g.columns:
            idx_bot         = g["Coverage_Pct"].idxmin()
            constraint_comp = g.loc[idx_bot, "Component"]
            constraint_desc = g.loc[idx_bot, "Component_Desc"] if "Component_Desc" in g.columns else ""
            coverage        = float(g["Coverage_Pct"].min())
            buildable       = float(max(0.0, g["Buildable_From_Comp"].min()))
            shortage        = float(g["Shortage"].sum())
        else:
            constraint_comp = None
            constraint_desc = None
            coverage        = 100.0
            buildable       = fg_qty
            shortage        = 0.0

        return pd.Series({
            "FG_Scheduled_Qty":      fg_qty,
            "Buildable_Qty":         buildable,
            "Material_Coverage_Pct": round(coverage, 1),
            "Constraint_Component":  constraint_comp,
            "Constraint_Desc":       constraint_desc,
            "Num_BOM_Components":    int(len(g)),
            "Total_Shortage":        shortage,
        })

    summary = enriched_df.groupby(grp_keys, sort=True).apply(
        _agg, include_groups=False
    ).reset_index()

    summary = summary.rename(columns={"FG": "Part_Number", "FG_Desc": "FG_Description"})

    # Add cost columns
    if cost_dated_lookup and "Part_Number" in summary.columns:
        pn = summary["Part_Number"].astype(str).str.strip()
        summary["Std_Cost"]           = pn.map(lambda p: cost_dated_lookup.get(p, {}).get("cost", 0.0))
        summary["Cost_Effective_Date"] = pn.map(lambda p: cost_dated_lookup.get(p, {}).get("date", ""))
        summary["Total_Cost"]         = summary["Buildable_Qty"] * summary["Std_Cost"]

    return summary


# ── 6. Shortage analysis ──────────────────────────────────────────────────────

def shortage_report(
    enriched_df: pd.DataFrame,
    week_col: str = "ProductionWeek",
) -> pd.DataFrame:
    """
    Returns a shortage summary: Component × Week → Shortage qty,
    sorted by total shortage descending.
    """
    if "Shortage" not in enriched_df.columns or enriched_df.empty:
        return pd.DataFrame()

    pvt = enriched_df.pivot_table(
        index=["Component", "Component_Desc"],
        columns=week_col,
        values="Shortage",
        aggfunc="sum",
        fill_value=0,
    )
    pvt["Total Shortage"] = pvt.sum(axis=1)
    pvt = pvt[pvt["Total Shortage"] > 0].sort_values("Total Shortage", ascending=False)
    return pvt.reset_index()


# ── 7. Full pipeline ──────────────────────────────────────────────────────────

def run_build_analysis(
    scheduled_df:  pd.DataFrame,
    bom_df:        pd.DataFrame | None,
    cost_df:       pd.DataFrame | None,
    stock_lookup:  dict[str, float] | None = None,
    fg_col:        str = "Product code",
    qty_col:       str = "ScheduledQty",
    week_col:      str = "ProductionWeek",
    week_date_col: str = "ProductionWeekDate",
    alt_bom:       int = 1,
    bom_level:     int = 1,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict]:
    """
    Full build analysis pipeline.

    Returns:
      bom_clean        – parsed and filtered BOM
      exploded         – component demand per FG per week (enriched with stock if provided)
      capability_df    – dashboard-ready capability summary per FG × week
      cost_dated_lookup – {PN: {"cost", "date"}} for traceability
    """
    cost_dated_lookup = build_cost_lookup_dated(cost_df)
    bom_clean         = parse_bom(bom_df, alt_bom=alt_bom, bom_level=bom_level)

    exploded = explode_bom_demand(
        scheduled_df, bom_clean,
        fg_col=fg_col, qty_col=qty_col,
        week_col=week_col, week_date_col=week_date_col,
    )

    if exploded.empty:
        return bom_clean, exploded, pd.DataFrame(), cost_dated_lookup

    has_stock = bool(stock_lookup)
    if has_stock:
        exploded = compute_build_capability(exploded, stock_lookup)

    capability = build_capability_summary(
        exploded, cost_dated_lookup,
        has_stock=has_stock,
        week_col=week_col, week_date_col=week_date_col,
    )

    return bom_clean, exploded, capability, cost_dated_lookup


# ── 8. Raw Material Coverage report ──────────────────────────────────────────

def build_rm_coverage_table(
    exploded_df: pd.DataFrame,
    stock_df: pd.DataFrame | None = None,
    rm_po_df: pd.DataFrame | None = None,
    week_col: str = "ProductionWeek",
    week_date_col: str = "ProductionWeekDate",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Build the Raw Material Coverage report.

    Returns:
      coverage_df – one row per component × metric (Demand / Receipts / Ending Balance)
                    Columns: Component, Component_Desc, UOM, Initial_Inventory,
                             Metric, <week1>, <week2>, ...
      kpi_df      – per-component KPIs:
                    Component, Description, Initial_Inventory,
                    First_Shortage_Week, Pct_Weeks_Covered, Total_Shortage_Qty

    Ending Balance formula:
      EB[week0] = Initial_Inventory + Receipts[week0] - Demand[week0]
      EB[weekN] = EB[weekN-1] + Receipts[weekN] - Demand[weekN]
    """
    if exploded_df is None or exploded_df.empty or week_col not in exploded_df.columns:
        return pd.DataFrame(), pd.DataFrame()

    # ── 1. Ordered week list ───────────────────────────────────────────────────
    if week_date_col in exploded_df.columns:
        week_map = (
            exploded_df[[week_col, week_date_col]]
            .dropna(subset=[week_col])
            .drop_duplicates()
            .sort_values(week_date_col)
        )
        weeks = list(week_map[week_col])
        week_dates_lookup: dict = dict(zip(week_map[week_col], week_map[week_date_col]))
    else:
        weeks = sorted(exploded_df[week_col].dropna().unique().tolist())
        week_dates_lookup = {}

    if not weeks:
        return pd.DataFrame(), pd.DataFrame()

    # ── 2. Demand pivot: Component × week ─────────────────────────────────────
    demand_pvt = exploded_df.pivot_table(
        index=["Component", "Component_Desc"],
        columns=week_col,
        values="Comp_Demand",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()

    # ── 3. Initial inventory ───────────────────────────────────────────────────
    stock_lookup = _build_stock_lookup(stock_df) if stock_df is not None else {}

    # UOM from stock_df if available
    uom_lookup: dict[str, str] = {}
    if stock_df is not None and not stock_df.empty:
        sdf = stock_df.copy()
        sdf.columns = [str(c).strip() for c in sdf.columns]
        comp_col_s = _col(sdf, ["Component", "Material", "PN", "Part Number", "Part"])
        uom_col_s = _col(sdf, ["UOM", "Unit of Measure", "Unit", "Base UOM", "Base Unit"])
        if comp_col_s and uom_col_s:
            for _, r in sdf.iterrows():
                c = str(r[comp_col_s]).strip()
                u = str(r[uom_col_s]).strip() if pd.notna(r.get(uom_col_s, "")) else ""
                if c:
                    uom_lookup[c] = u

    # ── 4. Weekly receipts from RM_PO ─────────────────────────────────────────
    # Map each PO Del/finish date to the production week that contains or
    # immediately follows that date (first week_date >= Del/finish).
    # This is the correct MRP receipt logic: material available from Del/finish
    # onward, so it feeds the first production week that falls on or after.
    receipts_map: dict[tuple, float] = {}  # (component, week_label) → qty
    if rm_po_df is not None and not rm_po_df.empty and week_dates_lookup:
        po_weekly = _build_weekly_po_table(rm_po_df)
        if not po_weekly.empty:
            sorted_weeks = sorted(
                week_dates_lookup.keys(),
                key=lambda w: pd.to_datetime(week_dates_lookup[w], errors="coerce"),
            )
            sorted_wdts = [
                pd.to_datetime(week_dates_lookup[w], errors="coerce").normalize()
                for w in sorted_weeks
            ]

            for _, po_row in po_weekly.iterrows():
                comp  = str(po_row["Component"]).strip()
                po_dt = pd.to_datetime(po_row["PO_WeekDate"], errors="coerce")
                if pd.isna(po_dt):
                    continue
                po_dt = po_dt.normalize()

                # First production week whose start date is >= Del/finish
                matched_wk = None
                for wk, wk_dt in zip(sorted_weeks, sorted_wdts):
                    if pd.notna(wk_dt) and wk_dt >= po_dt:
                        matched_wk = wk
                        break
                # If Del/finish is after the last production week, attach to last
                if matched_wk is None:
                    matched_wk = sorted_weeks[-1]

                key = (comp, matched_wk)
                receipts_map[key] = receipts_map.get(key, 0.0) + float(po_row["PO_Qty"])

    # ── 5. Build 3-row-per-component coverage table ───────────────────────────
    coverage_rows: list[dict] = []
    kpi_rows: list[dict] = []

    for _, comp_row in demand_pvt.iterrows():
        comp = str(comp_row["Component"]).strip()
        raw_desc = comp_row.get("Component_Desc", comp)
        desc = "" if str(raw_desc) in ("nan", "None", "", comp) else str(raw_desc).strip()
        initial_inv = stock_lookup.get(comp, 0.0)
        uom = uom_lookup.get(comp, "")

        row_d: dict = {"Component": comp, "Component_Desc": desc,
                       "UOM": uom, "Initial_Inventory": initial_inv, "Metric": "Demand"}
        row_r: dict = {"Component": "", "Component_Desc": "", "UOM": "",
                       "Initial_Inventory": None, "Metric": "Receipts"}
        row_b: dict = {"Component": "", "Component_Desc": "", "UOM": "",
                       "Initial_Inventory": None, "Metric": "Ending Balance"}

        balance = initial_inv
        first_shortage_wk = None
        shortage_count = 0
        net_deficit_weeks = 0.0   # sum of running negative balances (MRP urgency index)
        peak_shortage     = 0.0   # maximum single-point deficit (actual procurement gap)

        for wk in weeks:
            demand_val   = float(comp_row.get(wk, 0) or 0)
            receipts_val = receipts_map.get((comp, wk), 0.0)
            balance      = balance + receipts_val - demand_val

            row_d[wk] = demand_val
            row_r[wk] = receipts_val
            row_b[wk] = round(balance, 2)

            if balance < 0:
                if first_shortage_wk is None:
                    first_shortage_wk = wk
                shortage_count    += 1
                net_deficit_weeks += abs(balance)
                peak_shortage      = max(peak_shortage, abs(balance))

        coverage_rows.extend([row_d, row_r, row_b])

        pct_covered = round((len(weeks) - shortage_count) / len(weeks) * 100, 1) if weeks else 100.0
        kpi_rows.append({
            "Component":           comp,
            "Description":         desc,
            "Initial_Inventory":   initial_inv,
            "First_Shortage_Week": first_shortage_wk if first_shortage_wk is not None else "—",
            "Pct_Weeks_Covered":   pct_covered,
            "Peak_Shortage":       round(peak_shortage, 0),       # actionable: minimum qty to procure now
            "Net_Deficit_Weeks":   round(net_deficit_weeks, 0),   # urgency index (sum of negative balances)
        })

    coverage_df = pd.DataFrame(coverage_rows)
    kpi_df = pd.DataFrame(kpi_rows)
    return coverage_df, kpi_df


def rm_coverage_to_excel(
    coverage_df: pd.DataFrame,
    kpi_df: pd.DataFrame,
) -> bytes:
    """
    Returns a formatted xlsx file with two sheets:
      - 'RM Coverage'  : green header, red/green Ending Balance cells
      - 'KPI Summary'  : per-component KPI table
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = _io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "RM Coverage"

    HDR_FILL   = PatternFill("solid", fgColor="1B5E20")
    HDR_FONT   = Font(bold=True, color="FFFFFF")
    RED_FILL   = PatternFill("solid", fgColor="FFCDD2")
    RED_FONT   = Font(bold=True, color="B71C1C")
    GRN_FILL   = PatternFill("solid", fgColor="C8E6C9")
    GRN_FONT   = Font(color="1B5E20")
    ZERO_FILL  = PatternFill("solid", fgColor="F5F5F5")
    RCP_FILL   = PatternFill("solid", fgColor="E8F5E9")
    GRP_FILL   = PatternFill("solid", fgColor="DDEEFF")
    GRP_BORDER = Border(bottom=Side(style="medium", color="9E9E9E"))
    NUM_FMT    = "#,##0"
    RIGHT      = Alignment(horizontal="right")
    LEFT       = Alignment(horizontal="left")
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if coverage_df.empty:
        wb.save(buf)
        return buf.getvalue()

    info_cols = ["Component", "Component_Desc", "UOM", "Initial_Inventory", "Metric"]
    week_keys = [c for c in coverage_df.columns if c not in info_cols]

    headers = (
        ["Part Number", "Description", "UOM", "Initial Inventory", "Metric"]
        + [f"Wk {c}" if isinstance(c, int) else str(c) for c in week_keys]
    )
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
    ws.freeze_panes = "F2"

    for _, row in coverage_df.iterrows():
        values = [row.get(c, "") for c in info_cols + week_keys]
        ws.append(values)
        ri = ws.max_row
        metric = str(row.get("Metric", ""))

        # Info columns: light blue on Part Number row
        is_pn_row = bool(str(row.get("Component", "")).strip())
        for ci, col_key in enumerate(info_cols + week_keys, 1):
            cell = ws.cell(row=ri, column=ci)
            cell.alignment = RIGHT if ci > 4 else LEFT
            if is_pn_row and ci <= 5:
                cell.fill = GRP_FILL

            if col_key in week_keys:
                raw = row.get(col_key, "")
                try:
                    num = float(raw)
                except (ValueError, TypeError):
                    num = None

                if metric == "Ending Balance" and num is not None:
                    cell.number_format = NUM_FMT
                    if num < 0:
                        cell.fill = RED_FILL
                        cell.font = RED_FONT
                    elif num > 0:
                        cell.fill = GRN_FILL
                        cell.font = GRN_FONT
                    else:
                        cell.fill = ZERO_FILL
                elif metric == "Receipts":
                    cell.fill = RCP_FILL
                    cell.number_format = NUM_FMT
                elif metric == "Demand":
                    cell.number_format = NUM_FMT

        if metric == "Ending Balance":
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).border = GRP_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    for i in range(1, len(week_keys) + 1):
        ws.column_dimensions[get_column_letter(5 + i)].width = 11

    # KPI sheet
    if not kpi_df.empty:
        ws2 = wb.create_sheet("KPI Summary")
        ws2.append(list(kpi_df.columns))
        for ci in range(1, len(kpi_df.columns) + 1):
            cell = ws2.cell(row=1, column=ci)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
        for _, row in kpi_df.iterrows():
            ws2.append(list(row.values))

    wb.save(buf)
    return buf.getvalue()


# ── 9. Line × Week production output plan (CTB → FG Output) ──────────────────

def build_line_output_plan(
    result_df: pd.DataFrame,
    capability_df: pd.DataFrame | None = None,
    week_col: str = "ProductionWeek",
    qty_col: str = "ScheduledQty",
    line_col: str = "Line",
    line_name_col: str = "Line Name",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Line × Week production output plan.

    Planned  = sum(ScheduledQty) from result_df per Line × Week.
               This uses the already-capacity-constrained, std-pack-rounded schedule
               from run_query — the ground truth for what was planned.

    FG Output = Planned constrained by CTB (from capability_df).
               If capability_df is None or a line/week has no BOM match,
               FG Output defaults to Planned (no material constraint known).

    Shortage = Planned − FG Output

    Recovery_Week = first week after first shortage where Shortage == 0.
    """
    if result_df is None or result_df.empty or line_col not in result_df.columns:
        return pd.DataFrame(), pd.DataFrame()
    if week_col not in result_df.columns or qty_col not in result_df.columns:
        return pd.DataFrame(), pd.DataFrame()

    # ── 1. Planned from result_df (ScheduledQty with capacity + std pack) ────
    line_name_map: dict = {}
    if line_name_col in result_df.columns:
        line_name_map = (
            result_df.dropna(subset=[line_name_col])
            .groupby(line_col)[line_name_col]
            .first()
            .to_dict()
        )

    planned_agg = (
        result_df.groupby([line_col, week_col], sort=True)[qty_col]
        .sum()
        .reset_index()
        .rename(columns={line_col: "Line", qty_col: "Planned"})
    )

    weeks: list = sorted(planned_agg[week_col].unique())
    lines: list = sorted(planned_agg["Line"].unique())

    # ── 2. FG Output from capability_df (CTB-constrained buildable per Line × Week)
    buildable_agg: dict[tuple, float] = {}
    if capability_df is not None and not capability_df.empty and "Line" in capability_df.columns:
        bq_col = "Buildable_Qty" if "Buildable_Qty" in capability_df.columns else None
        pq_col = "FG_Scheduled_Qty" if "FG_Scheduled_Qty" in capability_df.columns else None
        if bq_col and pq_col:
            cap = capability_df.copy()
            cap["_buildable"] = pd.to_numeric(cap[bq_col], errors="coerce").fillna(0.0)
            cap["_planned_cap"] = pd.to_numeric(cap[pq_col], errors="coerce").fillna(0.0)
            cap["_buildable"] = cap[["_planned_cap", "_buildable"]].min(axis=1)
            for (line, wk), g in cap.groupby(["Line", week_col]):
                buildable_agg[(line, wk)] = float(g["_buildable"].sum())

    # ── 3. Build stacked output ───────────────────────────────────────────────
    planned_idx = planned_agg.set_index(["Line", week_col])["Planned"]

    agg_dict: dict[tuple, dict] = {
        (line, wk): {"Planned": float(planned_idx.get((line, wk), 0.0))}
        for line in lines for wk in weeks
    }
    for (line, wk), val in agg_dict.items():
        planned  = val["Planned"]
        if buildable_agg:
            # Use CTB-based buildable; cap at planned; default to planned when line/week missing
            fg_out = min(planned, buildable_agg.get((line, wk), planned))
        else:
            fg_out = planned
        val["FG_Output"] = fg_out
        val["Shortage"]  = max(0.0, planned - fg_out)

    plan_rows: list[dict] = []
    kpi_rows:  list[dict] = []

    for line in lines:
        lname = line_name_map.get(line, "")

        row_p: dict = {"Line": line, "Line_Name": lname, "Metric": "Planned"}
        row_o: dict = {"Line": line, "Line_Name": lname, "Metric": "FG Output"}
        row_s: dict = {"Line": line, "Line_Name": lname, "Metric": "Shortage"}

        first_shortage_wk = None
        recovery_wk       = None
        found_shortage    = False
        total_planned = total_output = total_shortage = 0.0

        for wk in weeks:
            vals     = agg_dict.get((line, wk), {"Planned": 0.0, "FG_Output": 0.0, "Shortage": 0.0})
            planned  = vals["Planned"]
            output   = vals["FG_Output"]
            shortage = vals["Shortage"]

            row_p[wk] = planned
            row_o[wk] = output
            row_s[wk] = shortage if shortage > 0 else 0.0

            total_planned  += planned
            total_output   += output
            total_shortage += shortage

            if shortage > 0:
                found_shortage = True
                if first_shortage_wk is None:
                    first_shortage_wk = wk
            elif found_shortage and shortage == 0 and recovery_wk is None:
                recovery_wk = wk

        plan_rows.extend([row_p, row_o, row_s])
        pct = round(total_output / total_planned * 100, 1) if total_planned > 0 else 100.0

        kpi_rows.append({
            "Line":                line,
            "Line_Name":           lname,
            "Total_Planned":       round(total_planned,  0),
            "Total_FG_Output":     round(total_output,   0),
            "Total_Shortage_Pcs":  round(total_shortage, 0),
            "Pct_Achievable":      pct,
            "First_Shortage_Week": first_shortage_wk if first_shortage_wk is not None else "—",
            "Recovery_Week":       (
                recovery_wk if recovery_wk is not None
                else ("No recovery yet" if found_shortage else "—")
            ),
        })

    plan_df = pd.DataFrame(plan_rows)
    kpi_df  = pd.DataFrame(kpi_rows)
    return plan_df, kpi_df


def line_output_to_excel(
    plan_df: pd.DataFrame,
    kpi_df:  pd.DataFrame,
) -> bytes:
    """
    Returns a formatted xlsx file with:
      - Sheet 'Line Output Plan'  : dark-green header, green FG Output rows, red Shortage rows
      - Sheet 'Line KPI'          : per-line KPI summary
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    buf = _io.BytesIO()
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Line Output Plan"

    HDR_FILL  = PatternFill("solid", fgColor="1B5E20")
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    GRN_FILL  = PatternFill("solid", fgColor="C8E6C9")
    GRN_FONT  = Font(bold=True, color="1B5E20")
    RED_FILL  = PatternFill("solid", fgColor="FFCDD2")
    RED_FONT  = Font(bold=True, color="B71C1C")
    PLN_FILL  = PatternFill("solid", fgColor="ECEFF1")
    LBL_FILL  = PatternFill("solid", fgColor="CFD8DC")
    NUM_FMT   = "#,##0"
    RIGHT     = Alignment(horizontal="right")
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left")
    BOT_BDR   = Border(bottom=Side(style="medium", color="9E9E9E"))

    if plan_df.empty:
        wb.save(buf)
        return buf.getvalue()

    info_cols = ["Line", "Line_Name", "Metric"]
    week_keys = [c for c in plan_df.columns if c not in info_cols]

    headers = ["Line", "Line Name", "Metric"] + [str(w) for w in week_keys]
    ws.append(headers)
    for ci in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
    ws.freeze_panes = "D2"

    prev_line = None
    for _, row in plan_df.iterrows():
        metric = str(row.get("Metric", ""))
        line   = str(row.get("Line", ""))

        values = [line, row.get("Line_Name", ""), metric] + [row.get(wk, 0) for wk in week_keys]
        ws.append(values)
        ri = ws.max_row

        # Suppress repeated Line/Line Name for rows 2-3 of each group
        if line == prev_line:
            ws.cell(row=ri, column=1).value = ""
            ws.cell(row=ri, column=2).value = ""
        prev_line = line

        # Choose fill/font by metric
        if metric == "FG Output":
            row_fill, row_font = GRN_FILL, GRN_FONT
        elif metric == "Shortage":
            row_fill, row_font = RED_FILL, RED_FONT
        else:
            row_fill, row_font = PLN_FILL, None

        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            if ci <= 3:
                cell.fill = LBL_FILL if ci == 3 else row_fill
                cell.alignment = LEFT
            else:
                cell.fill = row_fill
                cell.font = row_font
                cell.number_format = NUM_FMT
                cell.alignment = RIGHT
                # Zero shortage → blank for cleanliness
                if metric == "Shortage" and (cell.value or 0) == 0:
                    cell.value = None

        # Bottom border after Shortage row (last of each group)
        if metric == "Shortage":
            for ci in range(1, len(headers) + 1):
                ws.cell(row=ri, column=ci).border = BOT_BDR

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 14
    for i in range(1, len(week_keys) + 1):
        ws.column_dimensions[get_column_letter(3 + i)].width = 10

    # KPI sheet
    if not kpi_df.empty:
        ws2 = wb.create_sheet("Line KPI")
        ws2.append(list(kpi_df.columns))
        for ci in range(1, len(kpi_df.columns) + 1):
            cell = ws2.cell(row=1, column=ci)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
        for _, row in kpi_df.iterrows():
            ws2.append(list(row.values))
            ri = ws2.max_row
            pct = row.get("Pct_Achievable", 100.0)
            try:
                pct = float(pct)
            except (TypeError, ValueError):
                pct = 100.0
            color = "C8E6C9" if pct >= 95 else ("FFF9C4" if pct >= 80 else "FFCDD2")
            pct_col = list(kpi_df.columns).index("Pct_Achievable") + 1
            ws2.cell(row=ri, column=pct_col).fill = PatternFill("solid", fgColor=color)

    wb.save(buf)
    return buf.getvalue()
